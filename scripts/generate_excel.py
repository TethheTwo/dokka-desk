#!/usr/bin/env python3
"""Genera archivos Excel con openpyxl basado en un JSON de entrada.

Uso: python3 generate_excel.py <input.json> <output.xlsx>
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint as ChartDataPoint, SeriesLabel
from openpyxl.chart.label import DataLabelList


CHART_COLORS = [
    "2F7FD6",  # brand blue
    "16A34A",  # green
    "D97706",  # amber
    "DC2626",  # red
    "8B5CF6",  # violet
    "EC4899",  # pink
    "06B6D4",  # cyan
    "F97316",  # orange
    "14B8A6",  # teal
    "6366F1",  # indigo
    "84CC16",  # lime
    "A855F7",  # purple
]


def hex_to_rgb(hex_color: str) -> str:
    """Convierte hex sin # a formato RGB."""
    h = hex_color.lstrip("#")
    if len(h) == 6:
        return h
    return "2F7FD6"


def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_to_rgb(hex_color), end_color=hex_to_rgb(hex_color), fill_type="solid")


def make_font(bold=False, size=10, color=None, italic=False) -> Font:
    kwargs = {"bold": bold, "size": size, "italic": italic}
    if color:
        kwargs["color"] = hex_to_rgb(color)
    return Font(**kwargs)


THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def apply_cell(cell, value, font_kwargs=None, fill_color=None, alignment=None, border=True):
    cell.value = value
    if font_kwargs:
        cell.font = make_font(**font_kwargs)
    if fill_color:
        cell.fill = make_fill(fill_color)
    if alignment:
        cell.alignment = Alignment(**alignment)
    elif alignment is None and font_kwargs and font_kwargs.get("size", 10) > 12:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if border:
        cell.border = THIN_BORDER


def build_sheet(ws, sheet_def: dict, styles: dict):
    title = sheet_def.get("title")
    metadata = sheet_def.get("metadata")
    col_widths = sheet_def.get("col_widths", [])
    title_row_num = 1
    meta_row_num = 2
    num_cols = 4  # default

    # Title row
    if title:
        r = title_row_num
        ws.row_dimensions[r].height = 34
        cell = ws.cell(row=r, column=1)
        apply_cell(cell, title, font_kwargs={"bold": True, "size": 16, "color": styles["white"]},
                   fill_color=styles["brand"], alignment={"horizontal": "center", "vertical": "center"})

    # Metadata row
    meta_row = title_row_num
    if title:
        meta_row += 1
    if metadata:
        ws.row_dimensions[meta_row].height = 22
        cell = ws.cell(row=meta_row, column=1)
        apply_cell(cell, metadata, font_kwargs={"italic": True, "size": 10, "color": styles["text_muted"]},
                   fill_color=styles.get("bg_gray", "F1F5F9"),
                   alignment={"horizontal": "center", "vertical": "center"})
        meta_row += 1

    # Second metadata row
    metadata2 = sheet_def.get("metadata2")
    if metadata2:
        ws.row_dimensions[meta_row].height = 22
        cell = ws.cell(row=meta_row, column=1)
        apply_cell(cell, metadata2, font_kwargs={"italic": True, "size": 10, "color": styles["text_muted"]},
                   fill_color=styles.get("bg_gray", "F1F5F9"),
                   alignment={"horizontal": "center", "vertical": "center"})
        meta_row += 1

    # Determine if using blocks or flat table
    blocks = sheet_def.get("blocks")

    current_row = 1
    if title:
        current_row += 1
    if metadata:
        current_row += 1
    if metadata2:
        current_row += 1

    if blocks:
        # Compute num_cols for merging title/meta and sections
        max_cols = len(col_widths) if col_widths else 0
        for b in blocks:
            if b.get("type") == "table":
                hdrs = b.get("headers", [])
                if len(hdrs) > max_cols:
                    max_cols = len(hdrs)
        if title and max_cols > 1:
            ws.merge_cells(start_row=title_row_num, start_column=1,
                           end_row=title_row_num, end_column=max_cols)
        if metadata and max_cols > 1:
            ws.merge_cells(start_row=title_row_num + 1, start_column=1,
                           end_row=title_row_num + 1, end_column=max_cols)
        if metadata2 and max_cols > 1:
            ws.merge_cells(start_row=title_row_num + 2, start_column=1,
                           end_row=title_row_num + 2, end_column=max_cols)

        # Build from blocks (Dashboard Resumen style)
        for block in blocks:
            block_type = block.get("type", "blank")

            if block_type == "blank":
                h = block.get("height", 4)
                ws.row_dimensions[current_row].height = h
                current_row += 1

            elif block_type == "section":
                section_title = block.get("title", "")
                ws.row_dimensions[current_row].height = 24
                cell = ws.cell(row=current_row, column=1)
                apply_cell(cell, section_title,
                           font_kwargs={"bold": True, "size": 12, "color": styles["white"]},
                           fill_color=styles.get("brand_dark", "1F5A99"),
                           alignment={"vertical": "center"})
                # Merge all col widths
                num_cols = len(col_widths) if col_widths else 4
                if num_cols > 1:
                    ws.merge_cells(start_row=current_row, start_column=1,
                                   end_row=current_row, end_column=num_cols)
                current_row += 1

            elif block_type == "table":
                headers = block.get("headers", [])
                rows = block.get("rows", [])
                num_cols = max(len(headers), len(col_widths))

                # Header row
                ws.row_dimensions[current_row].height = block.get("header_height", 22)
                for i, h in enumerate(headers):
                    cell = ws.cell(row=current_row, column=i + 1)
                    apply_cell(cell, h,
                               font_kwargs={"bold": True, "size": 10, "color": styles["text"]},
                               fill_color=styles.get("bg_gray", "F1F5F9"),
                               alignment={"vertical": "center",
                                          "horizontal": "right" if i > 0 else "left"})
                current_row += 1

                # Data rows
                for ri, row_data in enumerate(rows):
                    zebra = ri % 2 == 1
                    ws.row_dimensions[current_row].height = block.get("data_height", 20)
                    for ci, val in enumerate(row_data):
                        cell = ws.cell(row=current_row, column=ci + 1)
                        apply_cell(cell, val,
                                   font_kwargs={"size": 10, "color": styles["text"]},
                                   fill_color=styles.get("bg_light") if zebra else styles.get("white", "FFFFFF"),
                                   alignment={"vertical": "center",
                                              "horizontal": "right" if ci > 0 else "left"})
                    current_row += 1
    else:
        # Flat table (Tickets List, AP, CG, Audit style)
        headers = sheet_def.get("headers", [])
        rows = sheet_def.get("rows", [])
        num_cols = max(len(headers), len(col_widths))

        if title and metadata:
            blank_row_num = meta_row_num + 1
            ws.row_dimensions[blank_row_num].height = 4
            current_row = max(current_row + 1, blank_row_num + 1)

        # Merge title across cols
        if title and num_cols > 1:
            ws.merge_cells(start_row=title_row_num, start_column=1,
                           end_row=title_row_num, end_column=num_cols)
        if metadata and num_cols > 1:
            ws.merge_cells(start_row=meta_row_num, start_column=1,
                           end_row=meta_row_num, end_column=num_cols)

        # Header row
        header_row = sheet_def.get("header_row", current_row)
        # If current_row doesn't match header_row, adjust
        if current_row < header_row:
            for _ in range(current_row, header_row):
                ws.row_dimensions[current_row].height = 4
                current_row += 1
        elif current_row > header_row:
            current_row = header_row

        ws.row_dimensions[current_row].height = sheet_def.get("header_row_height", 24)
        for i, h in enumerate(headers):
            cell = ws.cell(row=current_row, column=i + 1)
            apply_cell(cell, h,
                       font_kwargs={"bold": True, "size": 10, "color": styles["white"]},
                       fill_color=styles["brand"],
                       alignment={"horizontal": "center", "vertical": "center", "wrapText": True})
        current_row += 1

        # Data rows (with zebra striping)
        data_start = sheet_def.get("data_start_row", current_row)
        if current_row < data_start:
            current_row = data_start

        for ri, row_data in enumerate(rows):
            zebra = ri % 2 == 1
            data_row_height = sheet_def.get("data_row_height", 20)
            ws.row_dimensions[current_row].height = data_row_height
            for ci, val in enumerate(row_data):
                cell = ws.cell(row=current_row, column=ci + 1)
                apply_cell(cell, val,
                           font_kwargs={"size": 9, "color": styles["text"]},
                           fill_color=styles.get("bg_light") if zebra else styles.get("white", "FFFFFF"),
                           alignment={"vertical": "center", "wrapText": True})
            current_row += 1

    # Column widths
    for i, w in enumerate(col_widths):
        if w:
            ws.column_dimensions[get_column_letter(i + 1)].width = w


def build_charts(wb, chart_defs: list, sheets_map: dict):
    for ch in chart_defs:
        sheet_name = ch.get("sheet")
        ws = sheets_map.get(sheet_name)
        if ws is None:
            continue

        chart_type = ch.get("type", "bar")
        title = ch.get("title", "")
        data_ref = ch.get("data", {})
        cats_ref = ch.get("cats", {})
        pos = ch.get("position", {"col": 5, "row": 5})
        color = hex_to_rgb(ch.get("color", "2F7FD6"))
        y_title = ch.get("y_title", "")
        x_title = ch.get("x_title", "")

        if chart_type == "bar":
            chart = BarChart()
            chart.type = "bar"
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            continue

        chart.title = title

        if chart_type == "pie":
            pass
        elif chart_type == "bar":
            # Horizontal bars: x_title (values) on X-axis, y_title (categories) on Y-axis
            if x_title:
                chart.x_axis.title = x_title
            if y_title:
                chart.y_axis.title = y_title
        else:
            if x_title:
                chart.x_axis.title = x_title
            if y_title:
                chart.y_axis.title = y_title

        try:
            data = Reference(ws,
                             min_col=data_ref.get("min_col", 2),
                             min_row=data_ref.get("min_row", 1),
                             max_col=data_ref.get("max_col", 2),
                             max_row=data_ref.get("max_row", 1))
            cats = Reference(ws,
                             min_col=cats_ref.get("min_col", 1),
                             min_row=cats_ref.get("min_row", 1),
                             max_col=cats_ref.get("max_col", 1),
                             max_row=cats_ref.get("max_row", 1))
        except Exception:
            continue

        series_name = ch.get("series_name", title)

        # Number of data points
        num_pts = data_ref.get("max_row", 1) - data_ref.get("min_row", 1) + 1

        if chart_type == "pie":
            chart.add_data(data)
            chart.set_categories(cats)
            if chart.series and num_pts > 0:
                pts = []
                for pi in range(num_pts):
                    pt = ChartDataPoint(idx=pi)
                    pt.graphicalProperties.solidFill = hex_to_rgb(CHART_COLORS[pi % len(CHART_COLORS)])
                    pts.append(pt)
                chart.series[0].data_points = pts
        else:
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            if chart.series and num_pts > 0:
                chart.series[0].tx = SeriesLabel(v=series_name)
                chart.series[0].dLbls = DataLabelList(showVal=True)
                pts = []
                for pi in range(num_pts):
                    pt = ChartDataPoint(idx=pi)
                    pt.graphicalProperties.solidFill = hex_to_rgb(CHART_COLORS[pi % len(CHART_COLORS)])
                    pts.append(pt)
                chart.series[0].data_points = pts

        # Position
        col_letter = get_column_letter(pos.get("col", 5))
        row_num = pos.get("row", 5)
        anchor = f"{col_letter}{row_num}"

        # Shape size
        chart.width = ch.get("width", 18)
        chart.height = ch.get("height", 12)

        ws.add_chart(chart, anchor)


def main():
    if len(sys.argv) < 3:
        print("Uso: python3 generate_excel.py <input.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    with open(input_path, "r", encoding="utf-8") as f:
        payload = json.load(f)

    wb = Workbook()
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    styles = payload.get("styles", {})
    sheets_def = payload.get("sheets", [])
    charts_def = payload.get("charts", [])

    sheets_map = {}

    for i, sheet_def in enumerate(sheets_def):
        name = sheet_def.get("name", f"Sheet{i+1}")
        # Truncate to 31 chars (Excel limit)
        safe_name = name[:31]
        ws = wb.create_sheet(title=safe_name)
        sheets_map[name] = ws
        build_sheet(ws, sheet_def, styles)

    # Build charts (must be after all sheets are created)
    build_charts(wb, charts_def, sheets_map)

    # Creator metadata
    creator = payload.get("creator", "")
    if creator:
        wb.properties.creator = creator

    wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
